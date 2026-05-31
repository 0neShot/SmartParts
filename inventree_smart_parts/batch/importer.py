"""
Batch Importer
==============
Processes Excel/CSV files containing MPNs for bulk part creation.
Errors are collected per row – one failure never aborts the entire batch.
"""

import csv
import io
import logging
import threading
import uuid
from datetime import datetime
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, field

logger = logging.getLogger("inventree_smart_parts.batch")

# In-memory job store (persists for the lifetime of the worker process)
_batch_jobs: Dict[str, "BatchJob"] = {}


@dataclass
class BatchRowResult:
    """Result for a single row in the batch."""

    row_number: int = 0
    mpn: str = ""
    status: str = ""  # 'success', 'error', 'skipped', 'duplicate'
    action: str = ""  # 'created', 'updated', 'skipped'
    part_id: Optional[int] = None
    part_name: str = ""
    message: str = ""
    error: str = ""


@dataclass
class BatchJob:
    """Represents a batch import job."""

    job_id: str = ""
    status: str = "pending"  # 'pending', 'running', 'phase2', 'completed', 'failed'
    total_rows: int = 0
    processed_rows: int = 0
    created: int = 0
    updated: int = 0
    skipped: int = 0
    failed: int = 0
    started_at: Optional[str] = None
    completed_at: Optional[str] = None
    results: List[BatchRowResult] = field(default_factory=list)
    errors: List[Dict[str, Any]] = field(default_factory=list)
    # Assembly BOM fields (populated when create_assembly=True)
    assembly_name: str = ""
    assembly_part_id: Optional[int] = None
    assembly_url: str = ""
    assembly_bom_count: int = 0
    assembly_errors: List[str] = field(default_factory=list)

    @property
    def progress_percent(self) -> int:
        if self.total_rows == 0:
            return 0
        return int((self.processed_rows / self.total_rows) * 100)

    def to_dict(self) -> Dict[str, Any]:
        d = {
            "job_id": self.job_id,
            "status": self.status,
            "total_rows": self.total_rows,
            "processed_rows": self.processed_rows,
            "progress_percent": self.progress_percent,
            "created": self.created,
            "updated": self.updated,
            "skipped": self.skipped,
            "failed": self.failed,
            "started_at": self.started_at,
            "completed_at": self.completed_at,
            "errors": self.errors,
        }
        if self.assembly_name:
            d["assembly"] = {
                "name": self.assembly_name,
                "part_id": self.assembly_part_id,
                "url": self.assembly_url,
                "bom_count": self.assembly_bom_count,
                "errors": self.assembly_errors,
            }
        return d


def get_job(job_id: str) -> Optional[BatchJob]:
    """Get a batch job by ID."""
    return _batch_jobs.get(job_id)


def parse_upload_file(file_obj):
    """
    Parse an uploaded Excel or CSV file into a list of row dicts.

    Returns:
        Tuple of (rows, assembly_metadata) where:
          - rows is List[Dict[str, str]] with normalized column names
          - assembly_metadata is Dict[str, str] with keys 'name', 'description', 'revision'
            (populated for Altium CSVs; all empty-string for Excel/standard CSVs)
    """
    filename = getattr(file_obj, "name", "unknown")

    if filename.endswith((".xlsx", ".xls")):
        return _parse_excel(file_obj), {"name": "", "description": "", "revision": ""}
    elif filename.endswith(".csv"):
        return _parse_csv(file_obj)
    else:
        raise ValueError(f"Unsupported file format: {filename}. Use .xlsx or .csv")


def _parse_excel(file_obj) -> List[Dict[str, str]]:
    """Parse an Excel file using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel import. "
            "Add 'openpyxl' to plugins.txt and restart."
        )

    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)

    # First row = headers
    header_row = next(rows_iter, None)
    if not header_row:
        raise ValueError("Excel file is empty")

    headers = [str(h).strip() if h else "" for h in header_row]
    normalized_headers = _normalize_headers(headers)

    if "mpn" not in normalized_headers:
        raise ValueError(f"No 'MPN' column found. Available columns: {headers}")

    results = []
    for row_values in rows_iter:
        row_dict = {}
        for i, val in enumerate(row_values):
            if i < len(normalized_headers):
                key = normalized_headers[i]
                row_dict[key] = str(val).strip() if val else ""

        # Skip empty rows
        if row_dict.get("mpn", "").strip():
            results.append(row_dict)

    wb.close()
    return results


def _parse_csv(file_obj):
    """Parse a CSV file, with automatic Altium BOM detection.

    Uses the unified BomParser for Altium-format CSVs, which handles
    header sniffing, metadata interception, and row sanitisation in a
    single pass.

    Returns:
        Tuple of (rows, assembly_metadata).
    """
    content = file_obj.read()
    if isinstance(content, bytes):
        content = content.decode("utf-8-sig")  # Handle BOM

    reader = csv.DictReader(io.StringIO(content))

    if not reader.fieldnames:
        raise ValueError("CSV file is empty or has no headers")

    original_headers = list(reader.fieldnames)

    # ── Auto-detect Altium BOM format ──────────────────────────────
    from .altium_parser import is_altium_bom, BomParser

    if is_altium_bom(original_headers):
        # Unified parser: sniff columns, intercept metadata, filter rows
        parser = BomParser(original_headers)
        all_raw_rows = [dict(row) for row in reader]

        cm = parser.column_map
        logger.info(
            f"Altium BOM detected. "
            f"MPN col: '{cm.mpn_col}' "
            f"({'clean' if cm.mpn_clean else 'combined'}), "
            f"Mfr col: '{cm.mfr_col}', "
            f"Qty col: '{cm.qty_col}'"
        )

        bom_items, assembly_metadata = parser.parse_all(all_raw_rows)
        return bom_items, assembly_metadata

    # ── Fallback: generic CSV with normalised headers ──────────────
    assembly_metadata = {"name": "", "description": "", "revision": ""}
    results = []
    for row in reader:
        normalized_row = {}
        for h in original_headers:
            norm_key = _normalize_header(h)
            val = row.get(h, "")
            normalized_row[norm_key] = val.strip() if val else ""

        if normalized_row.get("mpn", "").strip():
            results.append(normalized_row)

    return results, assembly_metadata


def _normalize_headers(headers: List[str]) -> List[str]:
    """Normalize column headers to standard keys."""
    return [_normalize_header(h) for h in headers]


def _normalize_header(header: str) -> str:
    """Normalize a single header string."""
    h = header.lower().strip()
    mappings = {
        "mpn": "mpn",
        "manufacturer part number": "mpn",
        "herstellerteilenummer": "mpn",
        "teilenummer": "mpn",
        "part number": "mpn",
        "mfr part": "mpn",
        "manufacturer": "manufacturer",
        "hersteller": "manufacturer",
        "mfr": "manufacturer",
        "category": "category",
        "kategorie": "category",
        "description": "description",
        "beschreibung": "description",
        "quantity": "quantity",
        "menge": "quantity",
        "qty": "quantity",
        "anzahl": "quantity",
        "location": "location",
        "lagerort": "location",
        # Altium-specific columns
        "name": "name",
        "designator": "designator",
        "price": "price",
    }
    return mappings.get(h, h)


def process_batch(
    rows: List[Dict[str, str]],
    plugin_instance,
    assembly_name: str = "",
    assembly_description: str = "",
    assembly_revision: str = "",
    assembly_category_id: Optional[int] = None,
) -> str:
    """
    Start a batch import job in a background thread and return the job_id immediately.

    The job object is stored in _batch_jobs and updated in real-time by the
    worker thread so that batch_status() can return live progress to the frontend.

    When assembly_name is non-empty, Phase 2+3 are executed after all rows
    are processed: an Assembly Part is created and BomItems are linked.

    Args:
        rows:                 Parsed rows from the upload file
        plugin_instance:      Reference to the SmartPartsPlugin for settings access
        assembly_name:        If set, triggers assembly creation (Phase 2+3)
        assembly_description: Optional description for the assembly Part
        assembly_revision:    Optional revision string for the assembly Part

    Returns:
        job_id (8-char string) – available immediately, before processing starts
    """
    job_id = str(uuid.uuid4())[:8]
    job = BatchJob(
        job_id=job_id,
        total_rows=len(rows),
        status="pending",
        started_at=datetime.now().isoformat(),
        assembly_name=assembly_name.strip(),
    )
    _batch_jobs[job_id] = job

    # Start background worker thread so the HTTP response is returned immediately
    thread = threading.Thread(
        target=_run_batch,
        args=(
            job,
            rows,
            plugin_instance,
            assembly_description,
            assembly_revision,
            assembly_category_id,
        ),
        daemon=True,  # Dies with the process; won't block gunicorn shutdown
        name=f"smartparts-batch-{job_id}",
    )
    thread.start()

    logger.info(f"Batch job {job_id} queued ({len(rows)} rows) – thread started")
    return job_id


def _run_batch(
    job: "BatchJob",
    rows: List[Dict[str, str]],
    plugin_instance,
    assembly_description: str = "",
    assembly_revision: str = "",
    assembly_category_id: Optional[int] = None,
) -> None:
    """
    Worker function executed in a background thread.

    Phase 1: Process all CSV rows (search APIs, create/update parts).
    Phase 2+3: If assembly_name is set, create an Assembly Part and link BomItems.
    """
    from ..services.assembly_builder import create_assembly_with_bom, BomItemSpec

    job.status = "running"
    logger.info(f"Batch job {job.job_id} starting – {job.total_rows} rows")

    # Collect BOM specs for Phase 2+3
    bom_specs: List[BomItemSpec] = []

    # ── Phase 1: Import all rows ──────────────────────────────────────────────
    for idx, row in enumerate(rows, start=1):
        try:
            row_result = _process_single_row(row, idx, plugin_instance)
            job.results.append(row_result)

            if row_result.status == "success":
                if row_result.action == "created":
                    job.created += 1
                elif row_result.action == "updated":
                    job.updated += 1

                # Collect for BOM if the part exists
                if row_result.part_id and job.assembly_name:
                    try:
                        qty = float(row.get("quantity", 1) or 1)
                    except (ValueError, TypeError):
                        qty = 1.0
                    bom_specs.append(
                        BomItemSpec(
                            part_id=row_result.part_id,
                            quantity=qty,
                            mpn=row_result.mpn,
                            reference=row.get("designator", ""),
                        )
                    )

            elif row_result.status in ("skipped", "duplicate"):
                job.skipped += 1
                # Duplicates still have an existing part_id – include in BOM
                if row_result.part_id and job.assembly_name:
                    try:
                        qty = float(row.get("quantity", 1) or 1)
                    except (ValueError, TypeError):
                        qty = 1.0
                    bom_specs.append(
                        BomItemSpec(
                            part_id=row_result.part_id,
                            quantity=qty,
                            mpn=row_result.mpn,
                            reference=row.get("designator", ""),
                        )
                    )
            else:
                job.failed += 1
                job.errors.append(
                    {
                        "row": idx,
                        "mpn": row_result.mpn,
                        "error": row_result.error,
                    }
                )

        except Exception as e:
            logger.error(f"Batch row {idx} unexpected error: {e}", exc_info=True)
            job.failed += 1
            job.errors.append(
                {
                    "row": idx,
                    "mpn": row.get("mpn", "?"),
                    "error": str(e),
                }
            )
            job.results.append(
                BatchRowResult(
                    row_number=idx,
                    mpn=row.get("mpn", "?"),
                    status="error",
                    error=str(e),
                )
            )

        # Update progress counter after every row so the status endpoint reflects it
        job.processed_rows = idx

    # ── Phase 2+3: Assembly & BOM creation ───────────────────────────────────
    if job.assembly_name and bom_specs:
        job.status = "phase2"
        logger.info(
            f"Batch job {job.job_id} Phase 2+3: "
            f"creating assembly '{job.assembly_name}' with {len(bom_specs)} BOM items"
        )
        asm_result = create_assembly_with_bom(
            assembly_name=job.assembly_name,
            bom_specs=bom_specs,
            category_id=assembly_category_id,
            description=assembly_description,
            revision=assembly_revision,
        )
        job.assembly_part_id = asm_result.assembly_part_id
        job.assembly_url = asm_result.assembly_url
        job.assembly_bom_count = asm_result.bom_items_created
        job.assembly_errors = asm_result.errors
        if not asm_result.success:
            logger.error(f"Batch job {job.job_id} assembly failed: {asm_result.errors}")
    elif job.assembly_name and not bom_specs:
        job.assembly_errors = [
            "No parts were successfully imported – assembly not created."
        ]

    job.status = "completed"
    job.completed_at = datetime.now().isoformat()

    logger.info(
        f"Batch job {job.job_id} finished: "
        f"{job.created} created, {job.updated} updated, "
        f"{job.skipped} skipped, {job.failed} failed"
    )


def _process_single_row(
    row: Dict[str, str],
    row_number: int,
    plugin_instance,
) -> BatchRowResult:
    """
    Process a single batch row: search APIs → merge → check dups → create.
    """
    from ..api_clients import MouserClient, DigiKeyClient, LCSCClient
    from ..services.data_merger import merge_part_data
    from ..services.duplicate_checker import check_duplicate
    from ..services.category_mapper import fuzzy_match_category
    from ..services.part_creator import create_part_from_data

    mpn = row.get("mpn", "").strip()
    result = BatchRowResult(row_number=row_number, mpn=mpn)

    if not mpn:
        result.status = "skipped"
        result.message = "Empty MPN"
        return result

    # ── Step 1: Check duplicates ──
    dup_action = plugin_instance.get_setting("DUPLICATE_ACTION")
    dup_result = check_duplicate(mpn, row.get("manufacturer", ""))

    if dup_result.is_duplicate:
        if dup_action == "skip":
            result.status = "duplicate"
            result.message = f"Duplicate: exists as '{dup_result.existing_part_name}'"
            result.part_id = dup_result.existing_part_id
            return result
        elif dup_action == "ask":
            result.status = "duplicate"
            result.message = f"Duplicate found: '{dup_result.existing_part_name}' – skipped in batch mode"
            result.part_id = dup_result.existing_part_id
            return result

    # ── Step 2: Search APIs ──
    api_results = []
    priority_str = plugin_instance.get_setting("API_PRIORITY") or "mouser,digikey,lcsc"
    priority_order = [p.strip() for p in priority_str.split(",") if p.strip()]

    if plugin_instance.get_setting("MOUSER_ENABLED"):
        try:
            client = MouserClient(api_key=plugin_instance.get_setting("MOUSER_API_KEY"))
            r = client.search_by_mpn(mpn)
            if r:
                api_results.append(r)
        except Exception as e:
            logger.warning(f"Mouser search failed for {mpn}: {e}")

    if plugin_instance.get_setting("DIGIKEY_ENABLED"):
        try:
            client = DigiKeyClient(
                client_id=plugin_instance.get_setting("DIGIKEY_CLIENT_ID"),
                client_secret=plugin_instance.get_setting("DIGIKEY_CLIENT_SECRET"),
            )
            r = client.search_by_mpn(mpn)
            if r:
                api_results.append(r)
        except Exception as e:
            logger.warning(f"DigiKey search failed for {mpn}: {e}")

    if plugin_instance.get_setting("LCSC_ENABLED"):
        try:
            client = LCSCClient()
            r = client.search_by_mpn(mpn)
            if r:
                api_results.append(r)
        except Exception as e:
            logger.warning(f"LCSC search failed for {mpn}: {e}")

    if not api_results:
        result.status = "error"
        result.error = "No results from any API source"
        return result

    # ── Step 3: Merge data ──
    merged = merge_part_data(api_results, priority_order)
    if not merged:
        result.status = "error"
        result.error = "Data merge produced no result"
        return result

    # Override with CSV data if present
    if row.get("manufacturer"):
        merged.manufacturer = row["manufacturer"]
    if row.get("description"):
        merged.description = row["description"]

    # ── Step 4: Map category ──
    category_id = None
    cat_from_row = row.get("category", "")
    cat_source = cat_from_row or merged.category

    if cat_source:
        threshold = plugin_instance.get_setting("FUZZY_THRESHOLD") or 65
        default_cat = plugin_instance.get_setting("DEFAULT_CATEGORY") or ""
        user_synonyms = plugin_instance.get_setting("CATEGORY_SYNONYMS") or "{}"
        learned_mappings = (
            plugin_instance.get_setting("LEARNED_CATEGORY_MAPPINGS") or "{}"
        )
        cat_id, cat_path, score = fuzzy_match_category(
            cat_source,
            threshold,
            default_cat,
            user_synonyms_json=user_synonyms,
            learned_mappings_json=learned_mappings,
        )
        category_id = cat_id

    # ── Step 5: Create part ──
    update_existing = dup_action == "update" and dup_result.is_duplicate
    existing_id = dup_result.existing_part_id if update_existing else None

    creation = create_part_from_data(
        part_data=merged,
        category_id=category_id,
        update_existing=update_existing,
        existing_part_id=existing_id,
        auto_create_companies=plugin_instance.get_setting("AUTO_CREATE_MANUFACTURERS"),
    )

    result.status = "success" if creation.success else "error"
    result.action = creation.action
    result.part_id = creation.part_id
    result.part_name = creation.part_name
    result.message = creation.message
    if creation.errors:
        result.error = "; ".join(creation.errors)

    return result
